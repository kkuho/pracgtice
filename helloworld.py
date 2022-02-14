
import pyupbit
import time
import datetime
from openpyxl import load_workbook

with open("djqqlxm.txt") as f: # upbit login
    lines = f.readlines()
    key = lines[0].strip() #strip 공백 제거
    secret = lines[1].strip()
    upbit = pyupbit.Upbit(key, secret)

global coinlist, sortlist, record

tickers = pyupbit.get_tickers(fiat="KRW")
coinlist = []
record = []
buyflag = True

def get_tickers(tickers):

    curtime = datetime.datetime.now()
    
    for item in tickers:
        h_itemlists_prev = pyupbit.get_ohlcv(item, interval="minute60", count=10).drop_duplicates()
        h_itemlists = h_itemlists_prev.reset_index().sort_values(by="index", ascending=False) # ticker의 ohlcv 값을 index를 reset해서 내림차순으로 정렬
        #print(h_itemlists)

        if( (h_itemlists.head(1)['index'].item().hour)%24 == curtime.hour) : # 첫번째 기록의 now date기록 중에 시간만 추출해서 지금 시간과 같다면

            coinlist.append([ item, 'null', 0, 0, curtime, h_itemlists.iloc[1]['value'] ]) # 두번째 데이터를 기록하네?? 왜?? 암튼 ticker별로 총거래금액을 기록
            
        time.sleep(0.1)
    
    sortlist = sorted(coinlist, key=lambda x:x[5], reverse= True) # 거래금액이 큰 순으로 정렬

    return sortlist[0:11] # 10개 ticker만 return함

def get_target_price(ticker): # 변동성 돌파 구간 계산 

    df = pyupbit.get_ohlcv(ticker)
    yesterday = df.iloc[-2]

    today_open = yesterday['close']
    yesterday_high = yesterday['high']
    yesterday_low = yesterday['low']
    target = today_open + (yesterday_high - yesterday_low) * 0.5
    
    return target

def get_yesterday_ma5(ticker): # 5일 평균선 
    df = pyupbit.get_ohlcv(ticker)
    close = df['close']
    ma = close.rolling(5).mean()
    return ma[-2]

def buy_crypto_currency(ticker, baseprice): # 매수
    # krw = upbit.get_balance()*0.999 # 예수금 확인
    # money = krw/10
    buy = upbit.buy_market_order(ticker, baseprice) # 예수금의 10%로만 매수 진행
    return buy

def sell_crypto_currency(ticker): # 매도 
    unit_sell = upbit.get_balance(ticker)
    sell = upbit.sell_market_order(ticker, unit_sell)
    return sell
    
def write_trade(trade): # trade 정보를 엑셀로 기록하는 함수
    print(trade)
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[0]]
    row = []
    day = trade['created_at'].split('T')[0] # 날짜
    time_action = trade['created_at'].split('T')[1].split('+')[0]
    bal = upbit.get_balances

    row.append(day)
    row.append(time_action)
    row.append(trade['market'])

    if trade['side'] == 'ask': 
        row.append('매도')
        row.append(trade['volume'])
        row.append(trade['uuid'])
        
    else : 
        row.append('매수') 
        row.append(trade['price'])
        row.append(trade['uuid'])
     
   
    ws.append(row)
    wb.save('upbitRecord.xlsx')

def write_target(ticker_input, target_price, ma5, curtime, baseprice): # trade 정보를 엑셀로 기록하는 함수
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[1]]
    row = []

    row.append(ticker_input)
    row.append(target_price)
    row.append(ma5)
    row.append(curtime)
    row.append(baseprice)

    ws.append(row)
    wb.save('upbitRecord.xlsx')

def getgapsize(askprice):
    return round(askprice * 0.03, min_unit(askprice))
    
def min_unit(askprice):
    if askprice >= 2000000:
        return -3
    elif askprice >= 1000000:
        return -2
    elif askprice >= 500000:
        return -2
    elif askprice >= 100000:
        return -1
    elif askprice >= 10000:
        return -1
    elif askprice >= 1000:
        return 0.5
    elif askprice >= 100:
        return 0
    elif askprice >= 10:
        return 1
    elif askprice >= 0:
        return 2

def write_record(record): # 거미줄 매매 정보 
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[2]]
    row = []

    for i in record:
        leng = len(i)
        for b in range(leng):
            row.append(i[b])
        


    ws.append(row)
    wb.save('upbitRecord.xlsx')


krw = upbit.get_balance()
current_price = pyupbit.get_current_price("KRW-SNT")
target_price = get_target_price("KRW-SNT")
ma5 = get_yesterday_ma5("KRW-SNT")
curtime = datetime.datetime.now()
askprice = pyupbit.get_orderbook("KRW-SNT")['orderbook_units'][0]['ask_price']
myval = upbit.get_balances()

ticker = [["KRW-ETH", 12], ["KRW-XRP", 232]]

for i in ticker:
    a = i[0]
    print(a)


# record = [['911f8bcc-f94e-4d88-a661-8cb1b539814c', curtime, 1],['911f8bcc-f94e-4d88-a661-8cb1b539814c', curtime, 2]]



buyflag = False
if buyflag == False and upbit.get_order('911f8bcc-f94e-4d88-a661-8cb1b539814c')['state'] == 'done':
    print("OK")


gaptick = getgapsize(askprice)
print(gaptick)
print(upbit.get_order('911f8bcc-f94e-4d88-a661-8cb1b539814c')['price'])
print(type(upbit.get_order('911f8bcc-f94e-4d88-a661-8cb1b539814c')['price']))

for item in record:
    if buyflag == True:
        cancel = upbit.cancel_order(item[0])
        print(cancel)
    
    elif buyflag == False and upbit.get_order(item[0])['state'] == 'done':
        
        avaTicker = 'KRW-SNT'
        sellsubprice = float(upbit.get_order(item[0])['price']) + gaptick
        print(sellsubprice)
        sellsubamount = float(upbit.get_order(item[0])['volume'])
        print(sellsubamount)
        write_record(record)
        ret = upbit.sell_limit_order(avaTicker, sellsubprice, sellsubamount)
        print(ret)


      
#     else:
#         pass