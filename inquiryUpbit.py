
import pyupbit
import time
import datetime
from openpyxl import load_workbook

with open("djqqlxm.txt") as f: # upbit login
    lines = f.readlines()
    key = lines[0].strip() #strip 공백 제거
    secret = lines[1].strip()
    upbit = pyupbit.Upbit(key, secret)

def get_target_price(ticker): # 변동성 돌파 구간 계산 

    df = pyupbit.get_ohlcv(ticker)
    yesterday = df.iloc[-2]

    today_open = yesterday['close']
    yesterday_high = yesterday['high']
    yesterday_low = yesterday['low']
    target = today_open + (yesterday_high - yesterday_low) * 0.5
    
    return target

def buy_crypto_currency(ticker): # 매수
    krw = upbit.get_balance() - 1000 # 예수금 확인
    orderbook = pyupbit.get_orderbook(ticker)

    asks = orderbook['orderbook_units']
    sell_price = asks[0]['ask_price']
    unit = krw / float(sell_price) # float 은 실수형으로 만들어주는 함수
    buy = upbit.buy_market_order(ticker, krw) # 현재 시장가로 구매 가능한 수량을 구매 ! 주의!! 실행하는 순간 매수됨
    return buy

def sell_crypto_currency(ticker): # 매도 
    unit_sell = upbit.get_balance(ticker)
    sell = upbit.sell_market_order(ticker, unit_sell)
    return sell
    


def get_yesterday_ma5(ticker): # 5일 평균선 
    df = pyupbit.get_ohlcv(ticker)
    close = df['close']
    ma = close.rolling(5).mean()
    return ma[-2]

def write_trade(trade): # trade 정보를 엑셀로 기록하는 함수
    print(trade)
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[0]]
    row = []
    day = trade['created_at'].split('T')[0] # 날짜
    time_action = trade['created_at'].split('T')[1].split('+')[0]
    
    row.append(day)
    row.append(time_action)

    coinname = {'KRW-ETC':'이더리움클래식', 'KRW-XRP':'리플', 'KRW-ETH':'이더리움', 'KRW-BTC':'비트코인캐시', 'KRW-OMG':'오미세고', 'KRW-EOS':'이오스'}
    row.append(coinname[trade['market']])

    if trade['side'] == 'ask': 
        row.append('매도')
        row.append(trade['volume'])
        row.append(trade['uuid']) 
    else : 
        row.append('매수') 
        row.append(trade['price'])
        row.append(trade['uuid'])

    ws.append(row)
    wb.save('upbitRecord.xlsx')

def write_target(ticker_input, target_price, ma5, now): # trade 정보를 엑셀로 기록하는 함수
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[1]]
    row = []

    row.append(ticker_input)
    row.append(target_price)
    row.append(ma5)
    row.append(now)

    ws.append(row)
    wb.save('upbitRecord.xlsx')

now = datetime.datetime.now()
mid = datetime.datetime(now.year, now.month, now.day) + datetime.timedelta(hours=33) # 다음 날 9시를 구하는 함수
ticker_input = input("원하는 ticker를 입력하세요 : ")
ma5 = get_yesterday_ma5(ticker_input)
target_price = get_target_price(ticker_input) # 이 코드 실행할 때 target 가격을 계산
current_price = pyupbit.get_current_price(ticker_input)

write_target(ticker_input, target_price, ma5, now)

while True:
    try:
        now = datetime.datetime.now()
        current_price = pyupbit.get_current_price(ticker_input)
        krw = upbit.get_balance()
        #ma5 = get_yesterday_ma5(ticker_input)
        #target_price = get_target_price(ticker_input)
        if mid < now < mid + datetime.timedelta(seconds=10): # 9시에서 10초 내에 있을 때 9시로 간주함
            print("정각입니다!!")
            target_price = get_target_price(ticker_input)
            ma5 = get_yesterday_ma5(ticker_input)
            now = datetime.datetime.now()
            mid = datetime.datetime(now.year, now.month, now.day) + datetime.timedelta(hours=33)
            trade = sell_crypto_currency(ticker_input)
            write_trade(trade)
        
        elif (current_price > target_price) and (current_price > ma5) and (krw >1000):
            
            print("가즈아아아!~~~")
            trade = buy_crypto_currency(ticker_input)
            write_trade(trade)

        else:
            print(now, "|", "현재가 : " , current_price, "|", "목표가 : ", target_price, "|",  "5일선 평균가 : ", ma5)
        
 
    except:
        print("에러 발생")
        
    time.sleep(1)
    
        






