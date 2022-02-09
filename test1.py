
import pyupbit
import time
import datetime
from openpyxl import load_workbook

with open("djqqlxm.txt") as f: # upbit login
    lines = f.readlines()
    key = lines[0].strip() #strip 공백 제거
    secret = lines[1].strip()
    upbit = pyupbit.Upbit(key, secret)

global coinlist

tickers = pyupbit.get_tickers(fiat="KRW")
curtime = datetime.datetime.now()
coinlist = []

def get_tickers(tickers):

    curtime = datetime.datetime.now()
    
    for item in tickers :
        h_itemlists_prev = pyupbit.get_ohlcv(item, interval="minute10", count=10).drop_duplicates()
        h_itemlists = h_itemlists_prev.reset_index().sort_values(by="index", ascending=False) # ticker의 ohlcv 값을 index를 reset해서 내림차순으로 정렬
        #print(h_itemlists)

        if( (h_itemlists.head(1)['index'].item().hour)%24 == curtime.hour) : # 첫번째 기록의 now date기록 중에 시간만 추출해서 지금 시간과 같다면

            coinlist.append([ item, 'null', 0, 0, curtime, h_itemlists.iloc[1]['value'] ]) # 두번째 데이터를 기록하네?? 왜?? 암튼 ticker별로 총거래금액을 기록
            
        time.sleep(0.1)
    
    sortlist = sorted(coinlist, key=lambda x:x[5], reverse= True) # 거래금액이 큰 순으로 정렬

    return sortlist[1:11] # 10개 ticker만 return함


def get_target_price(ticker): # 변동성 돌파 구간 계산 

    df = pyupbit.get_ohlcv(ticker)
    yesterday = df.iloc[-2]

    today_open = yesterday['close']
    yesterday_high = yesterday['high']
    yesterday_low = yesterday['low']
    target = today_open + (yesterday_high - yesterday_low) * 0.5
    
    return target

def get_yesterday_ma5(ticker): # 5일 평균선 
    df = pyupbit.get_ohlcv(ticker)
    close = df['close']
    ma = close.rolling(5).mean()
    return ma[-2]

def buy_crypto_currency(ticker): # 매수
    krw = upbit.get_balance()*0.999 # 예수금 확인
    # orderbook = pyupbit.get_orderbook(ticker)
    # asks = orderbook['orderbook_units']
    # buy_price = asks[0]['ask_price']
    # unit = krw / float(buy_price) # float 은 실수형으로 만들어주는 함수
    buy = upbit.buy_market_order(ticker, krw) # 현재 시장가로 구매 가능한 수량을 구매 ! 주의!! 실행하는 순간 매수됨
    return buy

def sell_crypto_currency(ticker): # 매도 
    unit_sell = upbit.get_balance(ticker)
    sell = upbit.sell_market_order(ticker, unit_sell)
    return sell
    
def write_trade(trade): # trade 정보를 엑셀로 기록하는 함수
    print(trade)
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[0]]
    row = []
    day = trade['created_at'].split('T')[0] # 날짜
    time_action = trade['created_at'].split('T')[1].split('+')[0]
    bal = upbit.get_balances

    row.append(day)
    row.append(time_action)
    row.append(trade['market'])

    if trade['side'] == 'ask': 
        row.append('매도')
        row.append(trade['volume'])
        row.append(trade['uuid'])
        row.append(bal[1]['avg_buy_price'])
        row.append(bal[1]['balance']) 

    else : 
        row.append('매수') 
        row.append(trade['price'])
        row.append(trade['uuid'])
        row.append(bal[1]['avg_buy_price'])
        row.append(bal[1]['balance'])
        

    ws.append(row)
    wb.save('upbitRecord.xlsx')

def write_target(ticker_input, target_price, ma5, curtime): # trade 정보를 엑셀로 기록하는 함수
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[1]]
    row = []

    row.append(ticker_input)
    row.append(target_price)
    row.append(ma5)
    row.append(curtime)

    ws.append(row)
    wb.save('upbitRecord.xlsx')

curtime = datetime.datetime.now()
am9 = datetime.datetime.now() + datetime.timedelta(minutes=2) # 다음 날 9시를 구하는 함수
ticker_input = get_tickers(tickers)[0][0]
print(ticker_input)

ma5 = get_yesterday_ma5(ticker_input)
target_price = get_target_price(ticker_input) # 이 코드 실행할 때 target 가격을 계산
current_price = pyupbit.get_current_price(ticker_input)

write_target(ticker_input, target_price, ma5, curtime)

while True:
    try:
        curtime = datetime.datetime.now()
        current_price = pyupbit.get_current_price(ticker_input)
        krw = upbit.get_balance()
        myval = upbit.get_balances()
        bought_coin = False

        if am9 < curtime < am9 + datetime.timedelta(seconds=10): # 9시에서 10초 내에 있을 때 9시로 간주함
            print("Update!!")
            ticker_input = get_tickers(tickers)[0][0]
            target_price = get_target_price(ticker_input)
            ma5 = get_yesterday_ma5(ticker_input)
            curtime = datetime.datetime.now()
            am9 = datetime.datetime(curtime.year, curtime.month, curtime.day) + datetime.timedelta(hours=33)


        if (current_price > target_price) and (current_price > ma5) and (krw >1000):
            
            print("가즈아아아!~~~")
            trade = buy_crypto_currency(ticker_input)
            bought_coin = True

            write_trade(trade)
            write_target(ticker_input, target_price, ma5, curtime)

        # 현재가가 매수 평균가보다 3% 이상일 때 매도
        
        if bought_coin == True and (current_price > float(myval[1]['avg_buy_price']) * 1.03) :
            trade = sell_crypto_currency(ticker_input)
            bought_coin = False
            write_trade(trade)
        

        else:
            print(curtime, "|", "Ticker : ", ticker_input ,"| 현재가 : " , current_price, "|", "목표가 : ", target_price, "|",  "5일선 평균가 : ", ma5)
        
 
    except:
        print("에러 발생")
        
    time.sleep(1)
    
        






