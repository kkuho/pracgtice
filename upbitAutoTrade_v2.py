
#from email.mime import base
import pyupbit
import time
import datetime
from openpyxl import load_workbook
import pandas as pd

with open("djqqlxm.txt") as f: # upbit login
    lines = f.readlines()
    key = lines[0].strip() #strip 공백 제거
    secret = lines[1].strip()
    upbit = pyupbit.Upbit(key, secret)

global coinlist, sortlist, record

tickers = pyupbit.get_tickers(fiat="KRW")
coinlist = []
record = []
buyflag = True

def get_tickers(tickers):

    curtime = datetime.datetime.now()
    
    for item in tickers:
        h_itemlists_prev = pyupbit.get_ohlcv(item, interval="minute60", count=10).drop_duplicates()
        h_itemlists = h_itemlists_prev.reset_index().sort_values(by="index", ascending=False) # ticker의 ohlcv 값을 index를 reset해서 내림차순으로 정렬
        #print(h_itemlists)

        if( (h_itemlists.head(1)['index'].item().hour)%24 == curtime.hour) : # 첫번째 기록의 now date기록 중에 시간만 추출해서 지금 시간과 같다면

            coinlist.append([ item, 'null', 0, 0, curtime, h_itemlists.iloc[1]['value'] ]) # 두번째 데이터를 기록하네?? 왜?? 암튼 ticker별로 총거래금액을 기록
            
        time.sleep(0.1)
    
    sortlist = sorted(coinlist, key=lambda x:x[5], reverse= True) # 거래금액이 큰 순으로 정렬

    return sortlist[0:11] # 10개 ticker만 return함

def get_target_price(ticker): # 변동성 돌파 구간 계산 

    df = pyupbit.get_ohlcv(ticker)
    yesterday = df.iloc[-2]

    today_open = yesterday['close']
    yesterday_high = yesterday['high']
    yesterday_low = yesterday['low']
    target = today_open + (yesterday_high - yesterday_low) * 0.5
    
    return target

def get_yesterday_ma5(ticker): # 5일 평균선 
    df = pyupbit.get_ohlcv(ticker)
    close = df['close']
    ma = close.rolling(5).mean()
    return ma[-2]

def buy_crypto_currency(ticker, baseprice): # 매수
    # krw = upbit.get_balance()*0.999 # 예수금 확인
    # money = krw/10
    buy = upbit.buy_market_order(ticker, baseprice) # 예수금의 10%로만 매수 진행
    return buy

def sell_crypto_currency(ticker): # 매도 
    unit_sell = upbit.get_balance(ticker)
    sell = upbit.sell_market_order(ticker, unit_sell)
    return sell
    
def write_trade(trade, krw): # trade 정보를 엑셀로 기록하는 함수
    print(trade)
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[0]]
    row = []
    day = trade['created_at'].split('T')[0] # 날짜
    time_action = trade['created_at'].split('T')[1].split('+')[0]

    row.append(day)
    row.append(time_action)
    row.append(trade['market'])

    if trade['side'] == 'ask': 
        row.append('매도')
        row.append(trade['volume'])
        row.append(trade['uuid'])
        
    else : 
        row.append('매수') 
        row.append(trade['price'])
        row.append(trade['uuid'])
     
    row.append(krw)
    ws.append(row)
    wb.save('upbitRecord.xlsx')

def write_target(ticker_input, target_price, ma5, curtime, baseprice, krw): # trade 정보를 엑셀로 기록하는 함수
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[1]]
    row = []

    row.append(ticker_input)
    row.append(target_price)
    row.append(ma5)
    row.append(curtime)
    row.append(baseprice)
    row.append(krw)

    ws.append(row)
    wb.save('upbitRecord.xlsx')

def write_record(record): # 거미줄 매매 정보 
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[2]]
    row = []

    for a in record:
        row.clear()
        leng = len(a)
        for b in range(leng):
            row.append(a[b])

        ws.append(row)
        wb.save('upbitRecord.xlsx')

def write_balance(curtime, krw): # 계좌 balance를 입력하는 함수
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[3]]
    row = []

    row.append(curtime)
    row.append(krw)

    ws.append(row)
    wb.save('upbitRecord.xlsx')

def delete_trade(): # 거미줄 매수 매도 기록 지우는 함수
    
    wb = load_workbook('upbitRecord.xlsx')
    ws = wb[wb.sheetnames[2]]
    
    ws.delete_rows(2,1000)
    wb.save('upbitRecord.xlsx')

def check_record(): # 미체결 주문이 있는지 확인하고 있으면, 해당 uuid를 record에 저장
    
    # read_record = pd.read_excel("/Users/kuhojung/Documents/CodingWorkSpace/practice/upbitRecord.xlsx", sheet_name=2, usecols=['UUID', 'time', 'ticker', 'subgetprice', 'subgetamount', '매수 / 매도'])
    # record = read_record.values.tolist()

    myval = upbit.get_balances()

    if len(myval) > 1:
        preTicker = 'KRW-' + myval[1]['currency']
        pre_record = upbit.get_order(preTicker)
        gaptick = float(pre_record[1]['price']) - float(pre_record[0]['price'])
        print(gaptick)


        for i in pre_record:
            record.append([i['uuid'], i['market'], i['price'], i['side'], i['ord_type']])

    print(record)    
    return record
   
def getgapsize(askprice):
    return round(askprice * 0.03, min_unit(askprice))
    
def min_unit(askprice):
    if askprice >= 2000000:
        return -3
    elif askprice >= 1000000:
        return -2
    elif askprice >= 500000:
        return -2
    elif askprice >= 100000:
        return -1
    elif askprice >= 10000:
        return -1
    elif askprice >= 1000:
        return -1
    elif askprice >= 100:
        return 0
    elif askprice >= 10:
        return 1
    elif askprice >= 0:
        return 2

def rsi(ticker, count):
    data = pyupbit.get_ohlcv(ticker, interval='minute60')
    close_data = data['close']
    delta = close_data.diff()

    ups, downs = delta.copy(), delta.copy()
    ups[ups < 0] = 0
    downs[downs > 0] = 0

    au = ups.ewm(com=count-1, min_periods= count).mean()
    ad = downs.abs().ewm(com= count-1, min_periods= count).mean()

    rs = au / ad
    return pd.Series(100 - (100 / (1 + rs)), name = 'RSI')

def check_gaptick():
     
    myval = upbit.get_balances()

    if len(myval) > 1:
        preTicker = 'KRW-' + myval[1]['currency']
        pre_record = upbit.get_order(preTicker)
        leng = len(pre_record)
        gaptick = float(pre_record[leng-1]['price']) - float(pre_record[leng-2]['price'])
        print(gaptick)
    
    return gaptick

def check_baseprice():
    myval = upbit.get_balances()

    if len(myval) > 1:
        preTicker = 'KRW-' + myval[1]['currency']
        pre_record = upbit.get_order(preTicker)
        baseprice = pre_record[0]['locked']
        print(baseprice)
    
    return baseprice


record = check_record()
gaptick = check_gaptick()
baseprice = check_baseprice()


while True:
    # try :
    curtime = datetime.datetime.now()
    krw = upbit.get_balance()
    myval = upbit.get_balances()
    coinlist.clear()
    
    

    if len(myval) < 2 : # 보유한 coin이 있는지 확인. 2보다 작으면 보유한 coin이 없는 것으로 보고, 매수 로직 가동

        buyflag = True
        for item in tickers:
            
            h_itemlists_prev = pyupbit.get_ohlcv(item, interval="minute60", count=10).drop_duplicates()
            h_itemlists = h_itemlists_prev.reset_index().sort_values(by="index", ascending=False) 
            # ticker의 ohlcv 값을 index를 reset해서 내림차순으로 정렬

            if( (h_itemlists.head(1)['index'].item().hour)%24 == curtime.hour) : # 첫번째 기록의 now date기록 중에 시간만 추출해서 지금 시간과 같다면

                coinlist.append([ item, 'null', 0, 0, curtime, h_itemlists.iloc[1]['value'] ]) 
                # 두번째 데이터를 기록하네?? 왜?? 암튼 ticker별로 총거래금액을 기록

            sortlist = sorted(coinlist, key=lambda x:x[5], reverse= True)            
            time.sleep(0.2)
        

        for idx, ticker_input in enumerate(sortlist[0:13]) :

            current_price = pyupbit.get_current_price(ticker_input[0])
            target_price = get_target_price(ticker_input[0])
            ma5 = get_yesterday_ma5(ticker_input[0])
            curtime = datetime.datetime.now()
            askprice = pyupbit.get_orderbook(ticker_input[0])['orderbook_units'][0]['ask_price']
            rate_of_rise = round((current_price - target_price)/target_price * 100, 1)
            now_rsi = rsi(ticker_input[0], 14).iloc[-1]


            if (buyflag and current_price > target_price) and (current_price > ma5) and (krw >=5000) and (rate_of_rise <= 8):
                
                print("가즈아아아!~~~")
                
                write_balance(curtime, krw)
                baseprice = upbit.get_balance()//10 * 0.995 # 예수금의 10%를 baseprice로 매수 진행
                trade = buy_crypto_currency(ticker_input[0], baseprice)
                buyflag = False
                
                gaptick = getgapsize(askprice)
                record.clear()

                write_trade(trade, krw)
                write_target(ticker_input[0], target_price, ma5, curtime, baseprice, krw)

                for i in range(1,11):
                    subgetprice = 0
                    subgetprice = round(askprice - gaptick*i)

                    subamount = round(baseprice / subgetprice, 8 ) #지정가 구매 수량 정하기
                    ret = upbit.buy_limit_order(ticker_input[0], subgetprice, subamount) # 지정가 구매
                    print(ret) 
                    record.append([ret['uuid'], curtime, ticker_input[0], subgetprice, subamount, "거미줄매수"])
                    time.sleep(1)
                    
                write_record(record)
            

            else:
                print(curtime, "|", "Ticker : ", ticker_input[0] ,"| 현재가 : " , current_price, "|", "목표가 : ", target_price, "|",  "5일선 평균가 : ", ma5)
            time.sleep(0.5)
        
            
            
    else:
        
        avaTicker = 'KRW-' + myval[1]['currency']
        if pyupbit.get_current_price(avaTicker) > float(myval[1]['avg_buy_price']) * 1.02 :
                # 현재가가 매수 평균가보다 3% 이상일 때 매도
            
            for item in record:
                cancel = upbit.cancel_order(item[0])
                print(cancel)
                time.sleep(0.1)

            trade = sell_crypto_currency(avaTicker)
            print(trade)
            buyflag = True
            write_trade(trade, krw)

            record.clear()
            delete_trade()

        else : 
            pass

        for item in record[:]:
            
            if  upbit.get_order(item[0])['side'] == 'bid' and upbit.get_order(item[0])['state'] == 'done':
                avaTicker = 'KRW-' + myval[1]['currency']
                sellsubprice = float(upbit.get_order(item[0])['price']) + gaptick
                sellsubamount = round(baseprice / sellsubprice, 8 )
                ret = upbit.sell_limit_order(avaTicker, sellsubprice, sellsubamount)
                print(ret)
                record.append([ret['uuid'], datetime.datetime.now(), avaTicker, sellsubprice, sellsubamount, '추가 거미줄매도'])
                record.remove(item)
                delete_trade()
                write_record(record)

            elif upbit.get_order(item[0])['side'] == 'ask' and upbit.get_order(item[0])['state'] == 'done':
                avaTicker = 'KRW-' + myval[1]['currency']
                buysubprice = float(upbit.get_order(item[0])['price']) - gaptick
                buysubamount = round(baseprice / buysubprice, 8 )
                ret = upbit.buy_limit_order(avaTicker, buysubprice, buysubamount)
                print(ret)
                record.append([ret['uuid'], datetime.datetime.now(), avaTicker, buysubprice, buysubamount, '추가 거미줄매수'])
                record.remove(item)
                delete_trade()
                write_record(record)
            time.sleep(0.2)
                
            
        rate_of_return = round((pyupbit.get_current_price(avaTicker)-float(myval[1]['avg_buy_price'])) / float(myval[1]['avg_buy_price']) * 100 , 1)
        print(curtime, "|", "Ticker : ", avaTicker ,"| 현재가 : " , pyupbit.get_current_price(avaTicker), "| 평균매수가 : ", myval[1]['avg_buy_price'], " | 수익률 : ", rate_of_return, "%")
        
    
 
    # except:
    #     print("에러 발생")
        
    time.sleep(0.1)
    
        






